import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine, desc
from sqlalchemy.orm import sessionmaker

from models import (Event, EventClicked, EventDecision, PrequestionnaireAnswer,
                    SurveyAnswer, TrainingEvent, TrainingEventDecision, User)


def _to_dict(model):
    if model is None:
        return {}
    d = model.__dict__
    d.pop('_sa_instance_state', None)   # remove this sqlalchemy state variable
    return d


def _create_sheet_for_table(session, wb, sheet_name, model):
    ws = wb.create_sheet(sheet_name)
    results = [_to_dict(d) for d in session.query(model)]
    headers = list(results[0].keys())

    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])
    for row in range(2, len(results) + 2):
        for field in range(len(headers)):
            ws.cell(row=row, column=field+1,
                    value=results[row-2].get(headers[field], None))


def dump_db_to_excel(excel_dir, filename, pg_username, pg_password, pg_host, pg_database):
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
    excel_file = Path(
        excel_dir, f'{filename}_{datetime.now().strftime("%Y%m%d_%H-%M-%S")}.xlsx')

    engine = create_engine(
        f'postgresql+psycopg2://{pg_username}:{pg_password}@{pg_host}/{pg_database}')
    Session = sessionmaker(bind=engine)
    session = Session()

    wb = Workbook()
    wb.remove_sheet(wb.active)
    models = [User,
              PrequestionnaireAnswer,
              TrainingEvent,
              TrainingEventDecision,
              Event,
              EventClicked,
              EventDecision,
              SurveyAnswer]

    for m in models:
        _create_sheet_for_table(session, wb, m.__name__, m)

    wb.save(excel_file)
    session.close()


if __name__ == "__main__":
    dump_db_to_excel(excel_dir='excel',
                     filename='cry-wolf',
                     pg_username='postgres',
                     pg_password='postgres',
                     pg_host='localhost',
                     pg_database='crywolf')
